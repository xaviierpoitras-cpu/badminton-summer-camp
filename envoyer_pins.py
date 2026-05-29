"""
envoyer_pins.py
───────────────
Envoie un courriel personnalisé à chaque parent avec le PIN de leur enfant.
Chaque parent reçoit SEULEMENT le PIN de son enfant.

CONFIGURATION :
  1. Remplis les variables dans la section CONFIGURATION ci-dessous
  2. Lance : python envoyer_pins.py
  3. Tu verras un aperçu avant l'envoi réel
"""

import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
import time

# ═══════════════════════════════════════════════════════════
#  ⚙️  CONFIGURATION — remplis ces champs avant de lancer
# ═══════════════════════════════════════════════════════════

GMAIL_ADDRESS  = "badmintonexcellence@gmail.com"        # ton adresse Gmail
GMAIL_APP_PW   = "xxxx xxxx xxxx xxxx"apki frtv fycy tzju        # mot de passe d'application Gmail (16 car.)
CAMP_URL       = "https://xaviierpoitras-cpu.github.io/badminton-summer-camp/"
FICHIER_PINS   = "codes_acces_badminton.xlsx" # fichier Excel avec les PINs

# ═══════════════════════════════════════════════════════════

SUJET = "🏸 Badminton Excellence 2026 — Ton code d'accès"

def corps_email(nom_joueur, pin, groupe):
    return f"""\
Bonjour,

Voici le code d'accès personnel de {nom_joueur} pour la plateforme de places disponibles du camp Badminton Excellence 2026.

━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Joueur : {nom_joueur}
  Groupe : {groupe}
  Code PIN : {pin}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

Lien vers la plateforme :
{CAMP_URL}

Comment ça fonctionne :
1. Ouvre le lien ci-dessus
2. Entre le code PIN à 4 chiffres
3. Tu peux voir les places disponibles et faire des réservations

Guide d'utilisation complet :
{CAMP_URL}guide.html

⚠️  NE PARTAGE PAS ce code avec d'autres familles — il est unique à {nom_joueur}.

Si tu as des questions, réponds à ce courriel.

À bientôt sur le court !
L'équipe Badminton Excellence
"""

def lire_joueurs(fichier):
    wb = load_workbook(fichier)
    ws = wb.active
    joueurs = []
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2 header rows
        nom   = row[0]
        email = row[1]
        pin   = row[2]
        grp   = row[3]
        if nom and email and pin and isinstance(pin, (str, int)):
            joueurs.append({
                'nom':   str(nom).strip(),
                'email': str(email).strip(),
                'pin':   str(pin).zfill(4),
                'groupe': str(grp).strip() if grp else '',
            })
    return joueurs

def envoyer(joueurs, dry_run=True):
    if dry_run:
        print("=" * 60)
        print("  APERÇU — aucun courriel envoyé (dry run)")
        print("=" * 60)
        for j in joueurs:
            print(f"  → {j['email']:35} | {j['nom']:25} | PIN: {j['pin']}")
        print(f"\n  Total : {len(joueurs)} courriels")
        print("\nPour envoyer pour vrai, lance :")
        print("  python envoyer_pins.py --envoyer")
        return

    print(f"Envoi de {len(joueurs)} courriels...")
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PW)
        for i, j in enumerate(joueurs, 1):
            msg = MIMEMultipart("alternative")
            msg["Subject"] = SUJET
            msg["From"]    = GMAIL_ADDRESS
            msg["To"]      = j['email']
            msg.attach(MIMEText(corps_email(j['nom'], j['pin'], j['groupe']), "plain", "utf-8"))
            server.sendmail(GMAIL_ADDRESS, j['email'], msg.as_string())
            print(f"  [{i}/{len(joueurs)}] ✅ Envoyé → {j['email']} ({j['nom']})")
            time.sleep(0.5)  # évite les limites Gmail
    print(f"\n✅ Terminé — {len(joueurs)} courriels envoyés!")

if __name__ == "__main__":
    import sys
    joueurs = lire_joueurs(FICHIER_PINS)
    print(f"📋 {len(joueurs)} joueurs trouvés dans {FICHIER_PINS}")
    dry = "--envoyer" not in sys.argv
    envoyer(joueurs, dry_run=dry)
